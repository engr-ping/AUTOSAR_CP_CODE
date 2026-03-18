import os
import re
import json
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ArxmlToExcelConverter:
    def __init__(self, arxml_file: str, output_excel: str):
        self.arxml_file = arxml_file
        self.output_excel = output_excel
        self.tree = None
        self.ns = {'ns': 'http://autosar.org/schema/r4.0'}
        
        # 加载平台类型引用配置
        self._load_platform_type_references()
        
    def _load_platform_type_references(self):
        """加载平台类型引用配置"""
        config_file = 'basic_type_references.json'
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                self.platform_base_path = config.get('platform_base_path', '/DataTypeFromSignalOrGroup')
                self.platform_basic_types = config.get('basic_types', {})
                print(f"✅ 加载平台类型引用配置: {config_file}")
                print(f"  平台基础路径: {self.platform_base_path}")
                print(f"  支持的基本类型: {', '.join(self.platform_basic_types.keys())}")
            except Exception as e:
                print(f"⚠️ 加载平台类型引用配置失败: {e}")
                print("  使用默认配置")
                self.platform_base_path = '/DataTypeFromSignalOrGroup'
                self.platform_basic_types = {
                    'boolean': 'boolean',
                    'uint8': 'uint8', 'uint16': 'uint16', 'uint32': 'uint32', 'uint64': 'uint64',
                    'sint8': 'sint8', 'sint16': 'sint16', 'sint32': 'sint32', 'sint64': 'sint64',
                    'float32': 'float32', 'float64': 'float64'
                }
        else:
            print(f"⚠️ 平台类型引用配置文件不存在: {config_file}")
            print("  使用默认配置")
            self.platform_base_path = '/DataTypeFromSignalOrGroup'
            self.platform_basic_types = {
                'boolean': 'boolean',
                'uint8': 'uint8', 'uint16': 'uint16', 'uint32': 'uint32', 'uint64': 'uint64',
                'sint8': 'sint8', 'sint16': 'sint16', 'sint32': 'sint32', 'sint64': 'sint64',
                'float32': 'float32', 'float64': 'float64'
            }
        
    def _load_arxml(self):
        """加载 ARXML 文件"""
        print(f"🔍 加载 ARXML 文件: {self.arxml_file}")
        try:
            with open(self.arxml_file, 'rb') as f:
                self.tree = etree.parse(f)
            print("✅ ARXML 文件加载成功")
        except Exception as e:
            print(f"❌ 加载 ARXML 文件失败: {e}")
            raise
    
    def _get_element_text(self, element, tag_name):
        """获取命名空间元素的文本"""
        if element is not None:
            elem = element.find(f"{{{self.ns['ns']}}}{tag_name}")
            return elem.text if elem is not None else None
        return None
    
    def _get_type_reference(self, element):
        """从 TYPE-TREF 或 IMPLEMENTATION-DATA-TYPE-REF 获取类型引用"""
        # 尝试获取 TYPE-TREF 引用
        tref = element.find(".//ns:TYPE-TREF", self.ns)
        if tref is not None and tref.text:
            # 检查是否是平台引用
            path = tref.text
            # 提取最后一个部分作为类型名
            type_name = path.split('/')[-1]
            return type_name
        
        # 尝试获取 IMPLEMENTATION-DATA-TYPE-REF 引用
        idt_ref = element.find(".//ns:IMPLEMENTATION-DATA-TYPE-REF", self.ns)
        if idt_ref is not None and idt_ref.text:
            # 检查是否是平台引用
            path = idt_ref.text
            # 提取最后一个部分作为类型名
            type_name = path.split('/')[-1]
            return type_name
        
        return None
    
    def extract_types(self):
        """提取类型定义"""
        print("🔍 提取类型定义...")
        types_data = []
        
        # 提取 IMPLEMENTATION-DATA-TYPE（基本类型）
        primitive_types = self.tree.xpath("//ns:IMPLEMENTATION-DATA-TYPE", namespaces=self.ns)
        print(f"  - 找到 {len(primitive_types)} 个实现数据类型")
        
        # 按在文件中出现的顺序处理
        for prim in primitive_types:
            name = self._get_element_text(prim, "SHORT-NAME")
            category = self._get_element_text(prim, "CATEGORY")
            
            # 检查是否是基本类型
            if name and category and category in ['uint8', 'uint16', 'uint32', 'uint64', 'sint8', 'sint16', 'sint32', 'sint64', 'float32', 'float64', 'boolean', 'VALUE']:
                types_data.append([name, "basic", "", name])
            elif name and category == "STRUCTURE":
                # 处理结构体类型
                sub_elements = prim.find("ns:SUB-ELEMENTS", self.ns)
                if sub_elements is not None:
                    elements = sub_elements.findall("ns:IMPLEMENTATION-DATA-TYPE-ELEMENT", self.ns)
                    print(f"    - 结构体 {name} 有 {len(elements)} 个字段")
                    for i, elem in enumerate(elements):
                        field_name = self._get_element_text(elem, "SHORT-NAME")
                        field_type_ref = elem.find(".//ns:IMPLEMENTATION-DATA-TYPE-REF", self.ns)
                        field_type = field_type_ref.text.split('/')[-1] if field_type_ref is not None else "Unknown"
                        
                        if field_name and field_type:
                            if i == 0:
                                # 第一个字段，包含类型名和类别
                                types_data.append([name, "struct", field_name, field_type])
                            else:
                                # 后续字段，只添加字段名和类型
                                types_data.append(["", "", field_name, field_type])
            elif name and category == "ARRAY":
                # 处理数组类型
                sub_elements = prim.find("ns:SUB-ELEMENTS", self.ns)
                if sub_elements is not None:
                    elements = sub_elements.findall("ns:IMPLEMENTATION-DATA-TYPE-ELEMENT", self.ns)
                    for elem in elements:
                        array_size = self._get_element_text(elem, "ARRAY-SIZE")
                        if array_size:
                            # 获取数组元素类型
                            field_type_ref = elem.find(".//ns:IMPLEMENTATION-DATA-TYPE-REF", self.ns)
                            base_type = field_type_ref.text.split('/')[-1] if field_type_ref is not None else "Unknown"
                            array_def = f"{base_type}[{array_size}]"
                            types_data.append([name, "array", "", array_def])
                            print(f"    - 数组 {name} 包含 {array_size} 个 {base_type} 元素")
            elif name and category == "TYPE_REFERENCE":
                # 处理数组类型（备用）
                sub_elements = prim.find("ns:SUB-ELEMENTS", self.ns)
                if sub_elements is not None:
                    elements = sub_elements.findall("ns:IMPLEMENTATION-DATA-TYPE-ELEMENT", self.ns)
                    for elem in elements:
                        array_size = self._get_element_text(elem, "ARRAY-SIZE")
                        if array_size:
                            # 获取数组元素类型
                            field_type_ref = elem.find(".//ns:IMPLEMENTATION-DATA-TYPE-REF", self.ns)
                            base_type = field_type_ref.text.split('/')[-1] if field_type_ref is not None else "Unknown"
                            array_def = f"{base_type}[{array_size}]"
                            types_data.append([name, "array", "", array_def])
                            print(f"    - 数组 {name} 包含 {array_size} 个 {base_type} 元素")
        
        print(f"✅ 提取了 {len(types_data)} 条类型数据")
        return types_data
    
    def _extract_runnable_entities(self):
        """提取 RUNNABLE-ENTITY 信息 - 从 RUNNABLES 中获取可运行实体"""
        print("🔍 提取可运行实体信息...")
        runnable_entities = {}
        
        # 查找所有 RUNNABLE-ENTITY
        runnables = self.tree.xpath("//ns:RUNNABLE-ENTITY", namespaces=self.ns)
        print(f"  - 找到 {len(runnables)} 个可运行实体")
        
        for runnable in runnables:
            runnable_name = self._get_element_text(runnable, "SHORT-NAME")
            if not runnable_name:
                continue
            
            # 存储可运行实体的基本信息
            runnable_entities[runnable_name] = {
                'name': runnable_name,
                'ports': []
            }
            
            # 检查数据读取访问点 (DATA-READ-ACCESSS)
            read_access_points = runnable.xpath("ns:DATA-READ-ACCESSS", namespaces=self.ns)
            for point in read_access_points:
                var_accesses = point.xpath("ns:VARIABLE-ACCESS", namespaces=self.ns)
                for var_access in var_accesses:
                    access_name = self._get_element_text(var_access, "SHORT-NAME")
                    accessed_var = var_access.find("ns:ACCESSED-VARIABLE", self.ns)
                    if accessed_var is not None:
                        autosar_var = accessed_var.find("ns:AUTOSAR-VARIABLE-IREF", self.ns)
                        if autosar_var is not None:
                            port_ref = autosar_var.find("ns:PORT-PROTOTYPE-REF", self.ns)
                            target_ref = autosar_var.find("ns:TARGET-DATA-PROTOTYPE-REF", self.ns)
                            
                            if port_ref is not None and target_ref is not None:
                                port_name = port_ref.text.split('/')[-1]
                                interface_name = target_ref.text.split('/')[-1]
                                
                                # 将端口信息存储到对应的 runnable 中
                                runnable_entities[runnable_name]['ports'].append({
                                    'port_name': port_name,
                                    'interface_name': interface_name,
                                    'direction': 'R'  # 接收
                                })
            
            # 检查数据写入访问点 (DATA-WRITE-ACCESSS)
            write_access_points = runnable.xpath("ns:DATA-WRITE-ACCESSS", namespaces=self.ns)
            for point in write_access_points:
                var_accesses = point.xpath("ns:VARIABLE-ACCESS", namespaces=self.ns)
                for var_access in var_accesses:
                    access_name = self._get_element_text(var_access, "SHORT-NAME")
                    accessed_var = var_access.find("ns:ACCESSED-VARIABLE", self.ns)
                    if accessed_var is not None:
                        autosar_var = accessed_var.find("ns:AUTOSAR-VARIABLE-IREF", self.ns)
                        if autosar_var is not None:
                            port_ref = autosar_var.find("ns:PORT-PROTOTYPE-REF", self.ns)
                            target_ref = autosar_var.find("ns:TARGET-DATA-PROTOTYPE-REF", self.ns)
                            
                            if port_ref is not None and target_ref is not None:
                                port_name = port_ref.text.split('/')[-1]
                                interface_name = target_ref.text.split('/')[-1]
                                
                                # 将端口信息存储到对应的 runnable 中
                                runnable_entities[runnable_name]['ports'].append({
                                    'port_name': port_name,
                                    'interface_name': interface_name,
                                    'direction': 'S'  # 发送
                                })
            
            # 为了向后兼容，也检查旧的数据接收点和发送点
            receive_points = runnable.xpath("ns:DATA-RECEIVE-POINT-BY-ARGUMENTS", namespaces=self.ns)
            for point in receive_points:
                var_accesses = point.xpath("ns:VARIABLE-ACCESS", namespaces=self.ns)
                for var_access in var_accesses:
                    access_name = self._get_element_text(var_access, "SHORT-NAME")
                    accessed_var = var_access.find("ns:ACCESSED-VARIABLE", self.ns)
                    if accessed_var is not None:
                        autosar_var = accessed_var.find("ns:AUTOSAR-VARIABLE-IREF", self.ns)
                        if autosar_var is not None:
                            port_ref = autosar_var.find("ns:PORT-PROTOTYPE-REF", self.ns)
                            target_ref = autosar_var.find("ns:TARGET-DATA-PROTOTYPE-REF", self.ns)
                            
                            if port_ref is not None and target_ref is not None:
                                port_name = port_ref.text.split('/')[-1]
                                interface_name = target_ref.text.split('/')[-1]
                                
                                # 将端口信息存储到对应的 runnable 中
                                runnable_entities[runnable_name]['ports'].append({
                                    'port_name': port_name,
                                    'interface_name': interface_name,
                                    'direction': 'R'  # 接收
                                })
            
            send_points = runnable.xpath("ns:DATA-SEND-POINT-BY-ARGUMENTS", namespaces=self.ns)
            for point in send_points:
                var_accesses = point.xpath("ns:VARIABLE-ACCESS", namespaces=self.ns)
                for var_access in var_accesses:
                    access_name = self._get_element_text(var_access, "SHORT-NAME")
                    accessed_var = var_access.find("ns:ACCESSED-VARIABLE", self.ns)
                    if accessed_var is not None:
                        autosar_var = accessed_var.find("ns:AUTOSAR-VARIABLE-IREF", self.ns)
                        if autosar_var is not None:
                            port_ref = autosar_var.find("ns:PORT-PROTOTYPE-REF", self.ns)
                            target_ref = autosar_var.find("ns:TARGET-DATA-PROTOTYPE-REF", self.ns)
                            
                            if port_ref is not None and target_ref is not None:
                                port_name = port_ref.text.split('/')[-1]
                                interface_name = target_ref.text.split('/')[-1]
                                
                                # 将端口信息存储到对应的 runnable 中
                                runnable_entities[runnable_name]['ports'].append({
                                    'port_name': port_name,
                                    'interface_name': interface_name,
                                    'direction': 'S'  # 发送
                                })
        
        print(f"  - 提取了 {sum(len(runnable['ports']) for runnable in runnable_entities.values())} 个端口访问")
        return runnable_entities
    
    def _extract_events_and_tasks(self):
        """提取事件和任务信息 - 从 SWC-INTERNAL-BEHAVIOR 中提取事件信息"""
        print("🔍 提取事件和任务信息...")
        event_runnable_mapping = {}
        
        # 查找所有 SWC-INTERNAL-BEHAVIOR
        behaviors = self.tree.xpath("//ns:RUNNABLES", namespaces=self.ns)
        
        for behavior in behaviors:
            # 查找所有事件
            events = behavior.xpath(".//ns:RUNNABLE-ENTITY | .//ns:DATA-RECEIVE-EVENT | .//ns:OPERATION-INVOKED-EVENT", namespaces=self.ns)
            for event in events:
                event_name = self._get_element_text(event, "SHORT-NAME")
                event_runnable_mapping[event_name] = {
                    'runnable': event_name  # 默认使用事件名作为可运行实体名
                }

        return event_runnable_mapping
    
    def extract_apis(self):
        """提取 API 接口定义"""
        print("🔍 提取 API 接口定义...")
        
        # 提取可运行实体
        runnable_entities = self._extract_runnable_entities()
        print(f"  - 可运行实体包含 {len(runnable_entities)} 个实体")
        
        # 提取事件和任务映射
        event_runnable_mapping = self._extract_events_and_tasks()
        
        # 提取 SENDER-RECEIVER-INTERFACE（按在文件中出现的顺序）
        interfaces = self.tree.xpath("//ns:SENDER-RECEIVER-INTERFACE", namespaces=self.ns)
        print(f"  - 找到 {len(interfaces)} 个接口")
        
        # 创建接口映射（保持顺序）
        interface_order = []
        interface_types = {}
        for iface in interfaces:
            iface_name = self._get_element_text(iface, "SHORT-NAME")
            if not iface_name:
                continue
                
            data_elements = iface.find("ns:DATA-ELEMENTS", self.ns)
            if data_elements is not None:
                var_data = data_elements.find("ns:VARIABLE-DATA-PROTOTYPE", self.ns)
                if var_data is not None:
                    type_ref = self._get_type_reference(var_data)
                    if iface_name and type_ref:
                        interface_order.append(iface_name)
                        interface_types[iface_name] = type_ref
        
        # 通过端口信息补充 API 详情
        p_ports = self.tree.xpath("//ns:P-PORT-PROTOTYPE", namespaces=self.ns)
        r_ports = self.tree.xpath("//ns:R-PORT-PROTOTYPE", namespaces=self.ns)
        
        print(f"  - 找到 {len(p_ports)} 个提供端口 (P-PORT)")
        print(f"  - 找到 {len(r_ports)} 个需求端口 (R-PORT)")
        
        # 为端口创建映射
        port_mapping = {}
        
        # 处理提供端口 (P-PORT)
        for port in p_ports:
            port_name = self._get_element_text(port, "SHORT-NAME")
            comp_spec = port.find(".//ns:PROVIDED-COM-SPECS", self.ns)
            if comp_spec is not None:
                data_ref = comp_spec.find(".//ns:DATA-ELEMENT-REF", self.ns)
                if data_ref is not None and data_ref.text:
                    interface_name = data_ref.text.split('/')[-1]
                    swc_name = self._get_swc_name_for_port(port)
                    
                    # 提取初始值
                    init_value = ""
                    init_value_elem = comp_spec.find(".//ns:INIT-VALUE", self.ns)
                    if init_value_elem is not None:
                        # 检查数值规范
                        num_value = init_value_elem.find(".//ns:NUMERICAL-VALUE-SPECIFICATION", self.ns)
                        if num_value is not None:
                            value_elem = num_value.find("ns:VALUE", self.ns)
                            if value_elem is not None and value_elem.text:
                                init_value = value_elem.text
                        # 检查记录规范（暂时用空字符串表示复杂类型）
                        record_value = init_value_elem.find(".//ns:RECORD-VALUE-SPECIFICATION", self.ns)
                        if record_value is not None:
                            init_value = "{}"  # 表示结构体初始值
                        # 检查数组规范
                        array_value = init_value_elem.find(".//ns:ARRAY-VALUE-SPECIFICATION", self.ns)
                        if array_value is not None:
                            init_value = "[]"  # 表示数组初始值
                    
                    port_mapping[interface_name] = {
                        'port_name': port_name,
                        'direction': 'S',  # Sender
                        'swc': swc_name if swc_name else ' ',
                        'init_value': init_value
                    }
        
        # 处理需求端口 (R-PORT)
        for port in r_ports:
            port_name = self._get_element_text(port, "SHORT-NAME")
            comp_spec = port.find(".//ns:REQUIRED-COM-SPECS", self.ns)
            if comp_spec is not None:
                data_ref = comp_spec.find(".//ns:DATA-ELEMENT-REF", self.ns)
                if data_ref is not None and data_ref.text:
                    interface_name = data_ref.text.split('/')[-1]
                    swc_name = self._get_swc_name_for_port(port)
                    port_mapping[interface_name] = {
                        'port_name': port_name,
                        'direction': 'R',  # Receiver
                        'swc': swc_name if swc_name else ' '
                    }
        
        # 根据接口顺序生成 API 数据
        apis_data = []
        
        # 首先，为每个接口创建基本条目
        for interface_name in interface_order:
            type_ref = interface_types.get(interface_name, "UnknownType")
            
            # 从端口映射获取方向、端口名、SWC和初始值
            direction = " "
            port_name = " "
            swc = " "
            init_value = "0"  # 默认初始值
            task = " "  # 初始化task变量
            
            if interface_name in port_mapping:
                port_info = port_mapping[interface_name]
                direction = port_info['direction']
                port_name = port_info['port_name']
                swc = port_info['swc']
                init_value = port_info.get('init_value', '0')
            
            # 尝试从可运行实体获取任务信息
            for runnable_name, runnable_info in runnable_entities.items():
                for port_info in runnable_info['ports']:
                    if port_info['interface_name'] == interface_name:
                        # 检查是否有任务映射
                        if runnable_name in event_runnable_mapping:
                            task = event_runnable_mapping[runnable_name]['runnable']
                        break
                if task != " ":
                    break
            
            apis_data.append([
                interface_name,
                type_ref,
                direction,
                swc,
                task,
                init_value
            ])
        
        print(f"✅ 提取了 {len(apis_data)} 条 API 数据（包含 Task 信息）")
        return apis_data
    
    def _get_swc_name_for_port(self, port_element):
        """通过端口元素获取所属 SWC 名称"""
        # 从端口向上查找 SWC 组件
        parent = port_element.getparent()
        while parent is not None:
            if parent.tag.endswith('APPLICATION-SW-COMPONENT-TYPE'):
                swc_name = self._get_element_text(parent, "SHORT-NAME")
                return swc_name
            parent = parent.getparent()
        return None
    
    def create_excel(self):
        """创建 Excel 文件"""
        print("📝 创建 Excel 文件...")
        wb = Workbook()
        
        # 删除默认工作表
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 创建 Types 工作表
        types_ws = wb.create_sheet("Types")
        types_ws.append(["Name", "Category", "Field Name", "Definition"])
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, 5):
            cell = types_ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加类型数据
        types_data = self.extract_types()
        for row_data in types_data:
            types_ws.append(row_data)
        
        # 调整列宽
        for col in range(1, 5):
            types_ws.column_dimensions[get_column_letter(col)].width = 20
        
        # 创建 APIs 工作表
        apis_ws = wb.create_sheet("APIs")
        apis_ws.append(["API Name", "Type Reference", "Port Direction", "SWC Name", "Task", "Initial Value"])
        
        # 设置表头样式
        for col in range(1, 7):
            cell = apis_ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加 API 数据
        apis_data = self.extract_apis()
        for row_data in apis_data:
            apis_ws.append(row_data)
        
        # 调整列宽
        for col in range(1, 7):
            apis_ws.column_dimensions[get_column_letter(col)].width = 20
        
        # 保存文件
        wb.save(self.output_excel)
        print(f"✅ Excel 文件创建完成: {self.output_excel}")
    
    def convert(self):
        """执行转换"""
        print("🔄 开始转换 ARXML 到 Excel...")
        
        # 检查输入文件
        if not os.path.exists(self.arxml_file):
            print(f"❌ 输入文件不存在: {self.arxml_file}")
            return
        
        try:
            self._load_arxml()
            self.create_excel()
            print("✅ 转换完成！")
        except Exception as e:
            print(f"❌ 转换过程中发生错误: {e}")
            import traceback
            traceback.print_exc()


def convert_arxml_to_excel(arxml_file: str, excel_file: str):
    """将 ARXML 文件转换为 Excel 文件
    
    Args:
        arxml_file: 输入的 ARXML 文件路径
        excel_file: 输出的 Excel 文件路径
    """
    converter = ArxmlToExcelConverter(arxml_file, excel_file)
    converter.convert()


# === 主程序入口 ===
if __name__ == '__main__':
    import sys
    
    if len(sys.argv) == 3:
        # 使用命令行参数
        input_arxml = sys.argv[1]
        output_excel = sys.argv[2]
    else:
        # 使用默认值
        input_arxml = 'SwcOutput.arxml'  # 输入的 ARXML 文件
        output_excel = 'converted_from_arxml2.xlsx'  # 输出的 Excel 文件
    
    # 检查输入文件是否存在
    if not os.path.exists(input_arxml):
        print(f"❌ 输入文件不存在: {input_arxml}")
        print("请确保 ARXML 文件存在。")
        print("当前目录文件列表:")
        for f in os.listdir('.'):
            if f.endswith('.arxml'):
                print(f"  - {f}")
        exit(1)
    
    convert_arxml_to_excel(input_arxml, output_excel)
