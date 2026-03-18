import os
import re
import json
import uuid
import sys
from lxml import etree
from openpyxl import load_workbook
from typing import List, Dict, Set, Tuple
from datetime import datetime


class SwcTaskArxmlGenerator:
    NS = 'http://autosar.org/schema/r4.0'

    BASIC_TYPES = {
        'boolean': 'boolean',
        'Boolean': 'boolean',
        'bool': 'boolean',
        'uint8': 'uint8',
        'Uint8': 'uint8',
        'UInt8': 'uint8',
        'uint16': 'uint16',
        'Uint16': 'uint16',
        'UInt16': 'uint16',
        'uint32': 'uint32',
        'Uint32': 'uint32',
        'UInt32': 'uint32',
        'uint64': 'uint64',
        'Uint64': 'uint64',
        'UInt64': 'uint64',
        'sint8': 'sint8',
        'Sint8': 'sint8',
        'SInt8': 'sint8',
        'sint16': 'sint16',
        'Sint16': 'sint16',
        'SInt16': 'sint16',
        'sint32': 'sint32',
        'Sint32': 'sint32',
        'SInt32': 'sint32',
        'sint64': 'sint64',
        'Sint64': 'sint64',
        'SInt64': 'sint64',
        'float32': 'float32',
        'Float32': 'float32',
        'float64': 'float64',
        'Float64': 'float64',
    }

    ARRAY_PATTERN = re.compile(r'^(\w+)\[(\d+)\]$')

    def __init__(self, excel_file: str, output_arxml: str, types_sheet: str = 'Types', apis_sheet: str = 'APIs'):
        self.excel_file = excel_file
        self.output_arxml = output_arxml
        self.types_sheet = types_sheet
        self.apis_sheet = apis_sheet
        self.type_definitions = {}
        self.api_entries = []  # (api_name, type_ref, port_dir, swc_name, task, init_value)
        self.created_idt_types = set()
        self.interface_created = set()
        
        # 加载平台类型引用配置
        self._load_platform_type_references()

    def _load_platform_type_references(self):
        """加载平台类型引用配置"""
        config_file = 'basic_type_references.json'
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                self.platform_base_path = config.get('platform_base_path', '/DataTypeFromSignalOrGroup')
                self.platform_basic_types = config.get('basic_types', {})
                print(f"✅ 加载平台类型引用配置: {config_file}")
                print(f"  平台基础路径: {self.platform_base_path}")
                print(f"  支持的基本类型: {', '.join(self.platform_basic_types.keys())}")
            except Exception as e:
                print(f"⚠️ 加载平台类型引用配置失败: {e}")
                print("  使用默认配置")
                self.platform_base_path = '/DataTypeFromSignalOrGroup'
                self.platform_basic_types = {
                    'boolean': 'boolean',
                    'uint8': 'uint8', 'uint16': 'uint16', 'uint32': 'uint32', 'uint64': 'uint64',
                    'sint8': 'sint8', 'sint16': 'sint16', 'sint32': 'sint32', 'sint64': 'sint64',
                    'float32': 'float32', 'float64': 'float64'
                }
        else:
            print(f"⚠️ 平台类型引用配置文件不存在: {config_file}")
            print("  使用默认配置")
            self.platform_base_path = '/DataTypeFromSignalOrGroup'
            self.platform_basic_types = {
                'boolean': 'boolean',
                'uint8': 'uint8', 'uint16': 'uint16', 'uint32': 'uint32', 'uint64': 'uint64',
                'sint8': 'sint8', 'sint16': 'sint16', 'sint32': 'sint32', 'sint64': 'sint64',
                'float32': 'float32', 'float64': 'float64'
            }

    def _generate_uuid(self) -> str:
        """生成标准 UUID4 字符串（带连字符）"""
        return str(uuid.uuid4())

    def _read_type_definitions(self):
        wb = load_workbook(self.excel_file)
        ws = wb[self.types_sheet]
        
        # 检查是否有 Field Name 列（第3列）
        has_field_name_col = False
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            if len(row) >= 3 and row[2] is not None and str(row[2]).strip().lower() == 'field name':
                has_field_name_col = True
                break
        
        types = {}
        type_order = []  # 保持类型顺序
        current_type = None
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, cat, field_name, defn = None, None, None, None
            
            if len(row) >= 1:
                name = row[0]
            if len(row) >= 2:
                cat = row[1]
            if len(row) >= 3:
                field_name = row[2]
            if len(row) >= 4:
                defn = row[3]
            
            if name:
                current_type = str(name).strip()
                
                # 添加到类型顺序列表（如果还没有添加过）
                if current_type not in type_order:
                    type_order.append(current_type)
                
                if cat and str(cat).strip().lower() == "basic" and defn:
                    def_val = str(defn).strip()
                    # 不区分大小写检测自引用
                    if def_val.lower() == current_type.lower():
                        print(f"⚠️ 跳过自引用 basic 类型: {current_type}")
                        # 但仍然添加到类型定义中，因为需要在 ARXML 中创建
                        types[current_type] = (str(cat).strip().lower(), [])
                        continue
                
                types[current_type] = (str(cat).strip().lower(), [])
                
                # 如果是 struct 且有 Field Name 列，则收集字段定义
                if str(cat).strip().lower() == "struct" and has_field_name_col:
                    if field_name and defn:
                        types[current_type][1].append((str(field_name).strip(), str(defn).strip()))
                elif str(cat).strip().lower() == "array":
                    # 数组类型，只处理 Definition 列
                    if defn:
                        types[current_type][1].append(str(defn).strip())
            elif current_type and field_name and defn and has_field_name_col:
                # 继续添加结构体字段
                types[current_type][1].append((str(field_name).strip(), str(defn).strip()))
            elif current_type and defn and not has_field_name_col:
                # 保持向后兼容：没有 Field Name 列时，继续使用 Definition 列
                types[current_type][1].append(str(defn).strip())
        
        # 如果没有 Field Name 列，需要重新处理 struct 类型
        if not has_field_name_col:
            for tname, (cat, defs) in types.items():
                if cat == "struct":
                    # 重新格式化为 (field_name, field_type) 元素列表
                    new_defs = []
                    for i, defn in enumerate(defs):
                        # 支持 "FieldName: FieldType" 格式
                        field_parts = defn.split(':', 1) if isinstance(defn, str) else (defn, defn)
                        if isinstance(defn, str) and len(field_parts) == 2:
                            field_name = field_parts[0].strip()
                            field_type = field_parts[1].strip()
                        else:
                            # 保持向后兼容
                            field_name = f"Field{i+1}"
                            field_type = defn
                        new_defs.append((field_name, field_type))
                    types[tname] = (cat, new_defs)
        
        self.type_definitions = types
        self.type_order = type_order  # 保存类型顺序

    def _read_api_definitions(self):
        wb = load_workbook(self.excel_file)
        ws = wb[self.apis_sheet]
        entries = []
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True):
            if not row[0] or not row[1]:
                continue
            api_name = str(row[0]).strip()
            type_ref = str(row[1]).strip()
            port_dir = (str(row[2]).strip().upper() if row[2] else 'S')
            swc_name = str(row[3]).strip() if row[3] else 'DefaultSWC'
            task = str(row[4]).strip() if row[4] else 'DefaultTask'  # ✅ 修复：确保 task 有默认值
            init_value = str(row[5]).strip() if row[5] else '0'  # 默认初始值为0
            entries.append((api_name, type_ref, port_dir, swc_name, task, init_value))
        self.api_entries = entries

    # === IDT (Implementation) ===
    def _create_idt(self, name: str, category: str, sub_elements=None):
        e = etree.Element(f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE")
        e.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(e, f"{{{self.NS}}}SHORT-NAME").text = name
        
        # 设置类别
        if category in self.BASIC_TYPES:
            # 基本类型使用 VALUE 类别
            etree.SubElement(e, f"{{{self.NS}}}CATEGORY").text = "VALUE"
            # 为 VALUE 类别的 ImplementationDataType 添加 SwDataDefProps
            props = etree.SubElement(e, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
            variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
            cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
            # 添加 BASE-TYPE-REF 引用平台基本类型
            base_ref = etree.SubElement(cond, f"{{{self.NS}}}BASE-TYPE-REF")
            base_ref.set("DEST", "SW-BASE-TYPE")
            # 使用平台引用路径
            base_ref.text = f"{self.platform_base_path}/{category}"
        elif category == "struct":
            etree.SubElement(e, f"{{{self.NS}}}CATEGORY").text = "STRUCTURE"
            if sub_elements:
                se = etree.SubElement(e, f"{{{self.NS}}}SUB-ELEMENTS")
                for field_name, field_type in sub_elements:
                    elem = etree.SubElement(se, f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE-ELEMENT")
                    elem.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
                    etree.SubElement(elem, f"{{{self.NS}}}SHORT-NAME").text = field_name
                    etree.SubElement(elem, f"{{{self.NS}}}CATEGORY").text = "TYPE_REFERENCE"  # ✅ 标准格式
                    props = etree.SubElement(elem, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
                    variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
                    cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
                    # ✅ 改为 IMPLEMENTATION-DATA-TYPE-REF
                    ref = etree.SubElement(cond, f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE-REF")
                    ref.set("DEST", "IMPLEMENTATION-DATA-TYPE")
                    ref.text = f"/DataTypes/ImplementationDataTypes/{field_type}"
        return e

    def _create_array_idt(self, array_name, base_type, size):
        e = etree.Element(f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE")
        e.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(e, f"{{{self.NS}}}SHORT-NAME").text = array_name
        etree.SubElement(e, f"{{{self.NS}}}CATEGORY").text = "ARRAY"  # 数组类型使用 ARRAY 类别
        sub = etree.SubElement(e, f"{{{self.NS}}}SUB-ELEMENTS")
        elem = etree.SubElement(sub, f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE-ELEMENT")
        elem.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(elem, f"{{{self.NS}}}SHORT-NAME").text = "Element"
        etree.SubElement(elem, f"{{{self.NS}}}CATEGORY").text = "TYPE_REFERENCE"  # 数组元素使用 TYPE_REFERENCE 类别
        etree.SubElement(elem, f"{{{self.NS}}}ARRAY-SIZE").text = str(size)
        etree.SubElement(elem, f"{{{self.NS}}}ARRAY-SIZE-SEMANTICS").text = "FIXED-SIZE"
        props = etree.SubElement(elem, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
        variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
        cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
        # 对于 TYPE_REFERENCE 类别的数组元素，使用 IMPLEMENTATION-DATA-TYPE-REF 引用 ImplementationDataType
        idt_ref = etree.SubElement(cond, f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE-REF")
        idt_ref.set("DEST", "IMPLEMENTATION-DATA-TYPE")
        # 引用本地包中的基本类型
        local_pkg_name = self.platform_base_path.split('/')[-1] if self.platform_base_path else "DataTypeFromSignalOrGroup"
        idt_ref.text = f"/{local_pkg_name}/{base_type}"
        return e

    def _create_and_add_type(self, tname, pkg):
        if tname in self.created_idt_types:
            return
        
        # 检查是否是基本类型 - 基本类型也需要在本地包中创建
        if tname in self.BASIC_TYPES or tname == 'boolean' or tname in self.platform_basic_types:
            # 基本类型也需要在本地包中创建
            print(f"  - 创建基本类型 '{tname}' 在本地包中")
            pkg.append(self._create_idt(tname, tname))
            self.created_idt_types.add(tname)
            return
            
        if tname not in self.type_definitions:
            return
            
        cat, defs = self.type_definitions[tname]
        
        if cat == "basic":
            # 基本类型也需要在本地包中创建
            print(f"  - 创建基本类型 '{tname}' 在本地包中")
            pkg.append(self._create_idt(tname, tname))
            self.created_idt_types.add(tname)
        elif cat == "array":
            if defs and len(defs) > 0:
                first_def = defs[0] if isinstance(defs[0], str) else defs[0][1] if isinstance(defs[0], tuple) else defs[0]
                match = self.ARRAY_PATTERN.match(first_def)
                if match:
                    base = match.group(1)
                    # 检查数组元素类型是否是基本类型
                    if base in self.platform_basic_types:
                        print(f"  - 数组 '{tname}' 使用平台基本类型 '{base}' 作为元素")
                    else:
                        self._create_and_add_type(base, pkg)
                    pkg.append(self._create_array_idt(tname, base, int(match.group(2))))
                    self.created_idt_types.add(tname)
                else:
                    print(f"❌ 无法解析数组定义: {first_def}")
            else:
                print(f"❌ 数组类型 '{tname}' 缺少定义")
        elif cat == "struct":
            if defs:  # 检查是否有字段定义
                for field_def in defs:
                    field_type = field_def[1] if isinstance(field_def, tuple) else field_def
                    # 检查字段类型是否是基本类型
                    if field_type.strip() not in self.platform_basic_types:
                        self._create_and_add_type(field_type.strip(), pkg)
                pkg.append(self._create_idt(tname, "struct", defs))
                self.created_idt_types.add(tname)
            else:
                print(f"⚠️ 结构体类型 '{tname}' 没有字段定义")

    # === ADT (Application) ===
    def _create_application_primitive_type(self, name: str):
        prim = etree.Element(f"{{{self.NS}}}APPLICATION-PRIMITIVE-DATA-TYPE")
        prim.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(prim, f"{{{self.NS}}}SHORT-NAME").text = name
        etree.SubElement(prim, f"{{{self.NS}}}CATEGORY").text = "VALUE"
        props = etree.SubElement(prim, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
        variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
        cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
        etree.SubElement(cond, f"{{{self.NS}}}SW-CALIBRATION-ACCESS").text = "READ-WRITE"
        return prim

    def _create_application_array_type(self, array_name: str, base_type: str, size: int):
        arr = etree.Element(f"{{{self.NS}}}APPLICATION-ARRAY-DATA-TYPE")
        arr.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(arr, f"{{{self.NS}}}SHORT-NAME").text = array_name
        etree.SubElement(arr, f"{{{self.NS}}}CATEGORY").text = "ARRAY"
        props = etree.SubElement(arr, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
        variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
        cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
        etree.SubElement(cond, f"{{{self.NS}}}SW-CALIBRATION-ACCESS").text = "READ-WRITE"
        elem = etree.SubElement(arr, f"{{{self.NS}}}ELEMENT")
        elem.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(elem, f"{{{self.NS}}}SHORT-NAME").text = array_name
        etree.SubElement(elem, f"{{{self.NS}}}CATEGORY").text = "VALUE"
        tref = etree.SubElement(elem, f"{{{self.NS}}}TYPE-TREF")
        tref.set("DEST", "APPLICATION-PRIMITIVE-DATA-TYPE")
        tref.text = f"/DataTypes/ApplicationDataTypes/{base_type}"
        etree.SubElement(elem, f"{{{self.NS}}}ARRAY-SIZE-SEMANTICS").text = "FIXED-SIZE"
        etree.SubElement(elem, f"{{{self.NS}}}MAX-NUMBER-OF-ELEMENTS").text = str(size)
        return arr

    def _create_application_record_type(self, struct_name: str, fields: List[Tuple[str, str]]):
        rec = etree.Element(f"{{{self.NS}}}APPLICATION-RECORD-DATA-TYPE")
        rec.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(rec, f"{{{self.NS}}}SHORT-NAME").text = struct_name
        etree.SubElement(rec, f"{{{self.NS}}}CATEGORY").text = "STRUCTURE"
        props = etree.SubElement(rec, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
        variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
        cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
        etree.SubElement(cond, f"{{{self.NS}}}SW-CALIBRATION-ACCESS").text = "READ-WRITE"
        element_spec = etree.SubElement(rec, f"{{{self.NS}}}ELEMENTS")
        for field_name, field_type in fields:
            comp = etree.SubElement(element_spec, f"{{{self.NS}}}APPLICATION-RECORD-ELEMENT")
            comp.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
            etree.SubElement(comp, f"{{{self.NS}}}SHORT-NAME").text = field_name
            tref = etree.SubElement(comp, f"{{{self.NS}}}TYPE-TREF")
            if self._is_array_type(field_type):
                tref.set("DEST", "APPLICATION-ARRAY-DATA-TYPE")
            elif field_type in self.type_definitions and self.type_definitions[field_type][0] == "struct":
                tref.set("DEST", "APPLICATION-RECORD-DATA-TYPE")
            else:
                tref.set("DEST", "APPLICATION-PRIMITIVE-DATA-TYPE")
            tref.text = f"/DataTypes/ApplicationDataTypes/{field_type}"
        return rec

    def _is_array_type(self, type_name: str) -> bool:
        if type_name in self.type_definitions:
            cat, _ = self.type_definitions[type_name]
            return cat == "array"
        return False

    # === Package Utilities ===
    def _ensure_package(self, root, pkg_name):
        ar_pkgs = root.find(f"{{{self.NS}}}AR-PACKAGES")
        if ar_pkgs is None:
            ar_pkgs = etree.SubElement(root, f"{{{self.NS}}}AR-PACKAGES")
        for pkg in ar_pkgs.findall(f"{{{self.NS}}}AR-PACKAGE"):
            if pkg.find(f"{{{self.NS}}}SHORT-NAME").text == pkg_name:
                elems = pkg.find(f"{{{self.NS}}}ELEMENTS")
                if elems is None:
                    elems = etree.SubElement(pkg, f"{{{self.NS}}}ELEMENTS")
                return elems
        pkg = etree.SubElement(ar_pkgs, f"{{{self.NS}}}AR-PACKAGE")
        etree.SubElement(pkg, f"{{{self.NS}}}SHORT-NAME").text = pkg_name
        elems = etree.SubElement(pkg, f"{{{self.NS}}}ELEMENTS")
        return elems

    def _ensure_subpackage(self, ar_pkgs, parent_name: str, child_name: str):
        parent_pkg = None
        for pkg in ar_pkgs.findall(f"{{{self.NS}}}AR-PACKAGE"):
            if pkg.find(f"{{{self.NS}}}SHORT-NAME").text == parent_name:
                parent_pkg = pkg
                break
        if parent_pkg is None:
            parent_pkg = etree.SubElement(ar_pkgs, f"{{{self.NS}}}AR-PACKAGE")
            etree.SubElement(parent_pkg, f"{{{self.NS}}}SHORT-NAME").text = parent_name

        for child in parent_pkg.findall(f"{{{self.NS}}}AR-PACKAGE"):
            if child.find(f"{{{self.NS}}}SHORT-NAME").text == child_name:
                elems = child.find(f"{{{self.NS}}}ELEMENTS")
                if elems is None:
                    elems = etree.SubElement(child, f"{{{self.NS}}}ELEMENTS")
                return elems

        new_child = etree.SubElement(parent_pkg, f"{{{self.NS}}}AR-PACKAGE")
        etree.SubElement(new_child, f"{{{self.NS}}}SHORT-NAME").text = child_name
        elems = etree.SubElement(new_child, f"{{{self.NS}}}ELEMENTS")
        return elems

    # === Interface & Port ===
    def _create_interface_adt(self, iface_name: str, type_name: str):
        iface = etree.Element(f"{{{self.NS}}}SENDER-RECEIVER-INTERFACE")
        iface.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(iface, f"{{{self.NS}}}SHORT-NAME").text = iface_name
        data_elems = etree.SubElement(iface, f"{{{self.NS}}}DATA-ELEMENTS")
        de = etree.SubElement(data_elems, f"{{{self.NS}}}VARIABLE-DATA-PROTOTYPE")
        de.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(de, f"{{{self.NS}}}SHORT-NAME").text = iface_name
        tref = etree.SubElement(de, f"{{{self.NS}}}TYPE-TREF")
        if self._is_array_type(type_name):
            tref.set("DEST", "APPLICATION-ARRAY-DATA-TYPE")
        elif type_name in self.type_definitions and self.type_definitions[type_name][0] == "struct":
            tref.set("DEST", "APPLICATION-RECORD-DATA-TYPE")
        else:
            tref.set("DEST", "APPLICATION-PRIMITIVE-DATA-TYPE")
        tref.text = f"/DataTypes/ApplicationDataTypes/{type_name}"
        return iface

    def _create_interface_correct_format(self, iface_name: str, type_name: str):
        """根据 sample.arxml 格式创建接口"""
        iface = etree.Element(f"{{{self.NS}}}SENDER-RECEIVER-INTERFACE")
        iface.set("UUID", self._generate_uuid())
        etree.SubElement(iface, f"{{{self.NS}}}SHORT-NAME").text = iface_name
        etree.SubElement(iface, f"{{{self.NS}}}IS-SERVICE").text = "false"
        
        data_elems = etree.SubElement(iface, f"{{{self.NS}}}DATA-ELEMENTS")
        de = etree.SubElement(data_elems, f"{{{self.NS}}}VARIABLE-DATA-PROTOTYPE")
        de.set("UUID", self._generate_uuid())
        
        # 使用接口名作为元素名
        element_name = iface_name
        etree.SubElement(de, f"{{{self.NS}}}SHORT-NAME").text = element_name
        
        # 引用 IMPLEMENTATION-DATA-TYPE
        tref = etree.SubElement(de, f"{{{self.NS}}}TYPE-TREF")
        tref.set("DEST", "IMPLEMENTATION-DATA-TYPE")
        
        # 所有类型都使用本地包引用
        # 从 platform_base_path 获取包名
        local_pkg_name = self.platform_base_path.split('/')[-1] if self.platform_base_path else "DataTypeFromSignalOrGroup"
        tref.text = f"/{local_pkg_name}/{type_name}"
        print(f"  - 接口 {iface_name} 使用本地类型引用: {tref.text}")
        
        return iface

    def _create_port_correct_format(self, port_name, iface_name, direction):
        """根据 sample.arxml 格式创建端口"""
        tag = "R-PORT-PROTOTYPE" if direction == 'R' else "P-PORT-PROTOTYPE"
        port = etree.Element(f"{{{self.NS}}}{tag}")
        port.set("UUID", self._generate_uuid())
        etree.SubElement(port, f"{{{self.NS}}}SHORT-NAME").text = port_name
        
        # 创建通信规范
        com_specs_tag = "REQUIRED-COM-SPECS" if direction == 'R' else "PROVIDED-COM-SPECS"
        com_specs = etree.SubElement(port, f"{{{self.NS}}}{com_specs_tag}")
        
        if direction == 'S':  # P-PORT
            spec = etree.SubElement(com_specs, f"{{{self.NS}}}NONQUEUED-SENDER-COM-SPEC")
            deref = etree.SubElement(spec, f"{{{self.NS}}}DATA-ELEMENT-REF")
            deref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
            deref.text = f"/InterfaceTypes/{iface_name}/{iface_name}"
            etree.SubElement(spec, f"{{{self.NS}}}HANDLE-OUT-OF-RANGE").text = "NONE"
            
            # 添加初始值 - 根据类型创建正确的规范
            # 首先需要获取接口的类型和初始值
            type_ref = None
            init_val = "0"
            for entry in self.api_entries:
                if entry[0] == iface_name:
                    type_ref = entry[1]
                    init_val = entry[5] if len(entry) > 5 else "0"
                    break
            
            if type_ref:
                init_value = etree.SubElement(spec, f"{{{self.NS}}}INIT-VALUE")
                
                if type_ref in self.type_definitions:
                    cat, sub_defs = self.type_definitions[type_ref]
                    
                    if cat == "struct":
                        # 结构体类型使用 RecordValueSpecification
                        record_value = etree.SubElement(init_value, f"{{{self.NS}}}RECORD-VALUE-SPECIFICATION")
                        etree.SubElement(record_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{type_ref}_Init"
                        
                        # 为每个字段添加初始值
                        if sub_defs:
                            fields_elem = etree.SubElement(record_value, f"{{{self.NS}}}FIELDS")
                            for field_name, field_type in sub_defs:
                                # 根据字段类型设置初始值
                                if field_type in self.type_definitions and self.type_definitions[field_type][0] == "struct":
                                    # 嵌套结构体 - 使用RECORD-VALUE-SPECIFICATION
                                    nested_record = etree.SubElement(fields_elem, f"{{{self.NS}}}RECORD-VALUE-SPECIFICATION")
                                    etree.SubElement(nested_record, f"{{{self.NS}}}SHORT-LABEL").text = f"{field_name}_Init"
                                    # 为嵌套结构体的所有字段添加初始值为0
                                    nested_defs = self.type_definitions[field_type][1]
                                    if nested_defs:
                                        nested_fields = etree.SubElement(nested_record, f"{{{self.NS}}}FIELDS")
                                        for nested_field_name, nested_field_type in nested_defs:
                                            nested_field_elem = etree.SubElement(nested_fields, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                                            etree.SubElement(nested_field_elem, f"{{{self.NS}}}SHORT-LABEL").text = f"{nested_field_name}_Init"
                                            etree.SubElement(nested_field_elem, f"{{{self.NS}}}VALUE").text = "0"
                                elif field_type in self.BASIC_TYPES or field_type in ['boolean', 'uint8', 'uint16', 'uint32', 'uint64', 'sint8', 'sint16', 'sint32', 'sint64', 'float32', 'float64']:
                                    # 基本类型字段 - 直接使用NUMERICAL-VALUE-SPECIFICATION
                                    num_value = etree.SubElement(fields_elem, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                                    etree.SubElement(num_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{field_name}_Init"
                                    etree.SubElement(num_value, f"{{{self.NS}}}VALUE").text = init_val if init_val != "{}" else "0"
                                else:
                                    # 其他类型字段 - 直接使用NUMERICAL-VALUE-SPECIFICATION
                                    num_value = etree.SubElement(fields_elem, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                                    etree.SubElement(num_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{field_name}_Init"
                                    etree.SubElement(num_value, f"{{{self.NS}}}VALUE").text = "0"
                    
                    elif cat == "array":
                        # 数组类型使用 ArrayValueSpecification
                        array_value = etree.SubElement(init_value, f"{{{self.NS}}}ARRAY-VALUE-SPECIFICATION")
                        etree.SubElement(array_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{type_ref}_Init"
                        
                        # 为数组元素添加初始值
                        elements_elem = etree.SubElement(array_value, f"{{{self.NS}}}ELEMENTS")
                        
                        # 获取数组第一个定义来解析数组大小
                        array_size = None
                        element_type = None
                        
                        if sub_defs and len(sub_defs) > 0:
                            first_def = sub_defs[0] if isinstance(sub_defs[0], str) else sub_defs[0][1] if isinstance(sub_defs[0], tuple) else sub_defs[0]
                            
                            # 先尝试从sub_defs的定义中解析（如 'uint8[10]'）
                            array_match = self.ARRAY_PATTERN.match(first_def)
                            if array_match:
                                element_type = array_match.group(1)
                                array_size = int(array_match.group(2))
                        
                        # 如果没有从sub_defs获取，尝试从type_ref直接解析
                        if array_size is None:
                            array_match = self.ARRAY_PATTERN.match(type_ref)
                            if array_match:
                                element_type = array_match.group(1)
                                array_size = int(array_match.group(2))
                        
                        # 生成数组元素初始值
                        if array_size and element_type:
                            for i in range(array_size):
                                elem_value = etree.SubElement(elements_elem, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                                etree.SubElement(elem_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{element_type}_{i}_Init"
                                # 使用Excel中的初始值，如果是数字则使用，否则使用0
                                default_val = "0"
                                if init_val and init_val != "[]" and init_val.isdigit():
                                    default_val = init_val
                                etree.SubElement(elem_value, f"{{{self.NS}}}VALUE").text = default_val
                    
                    else:
                        # 基本类型使用 NumericalValueSpecification
                        num_value = etree.SubElement(init_value, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                        etree.SubElement(num_value, f"{{{self.NS}}}VALUE").text = init_val if init_val else "0"
                else:
                    # 不在type_definitions中的类型（如基本类型uint16, boolean）也使用NumericalValueSpecification
                    num_value = etree.SubElement(init_value, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                    etree.SubElement(num_value, f"{{{self.NS}}}VALUE").text = init_val if init_val else "0"
            
            # 添加接口引用
            iface_ref = etree.SubElement(port, f"{{{self.NS}}}PROVIDED-INTERFACE-TREF")
            iface_ref.set("DEST", "SENDER-RECEIVER-INTERFACE")
            iface_ref.text = f"/InterfaceTypes/{iface_name}"
        else:  # R-PORT
            spec = etree.SubElement(com_specs, f"{{{self.NS}}}NONQUEUED-RECEIVER-COM-SPEC")
            deref = etree.SubElement(spec, f"{{{self.NS}}}DATA-ELEMENT-REF")
            deref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
            deref.text = f"/InterfaceTypes/{iface_name}/{iface_name}"
            etree.SubElement(spec, f"{{{self.NS}}}HANDLE-OUT-OF-RANGE").text = "NONE"
            etree.SubElement(spec, f"{{{self.NS}}}USES-END-TO-END-PROTECTION").text = "false"
            etree.SubElement(spec, f"{{{self.NS}}}ALIVE-TIMEOUT").text = "0"
            
            # 添加初始值 - R-PORT 也需要 INIT-VALUE
            type_ref = None
            init_val = "0"
            for entry in self.api_entries:
                if entry[0] == iface_name:
                    type_ref = entry[1]
                    init_val = entry[5] if len(entry) > 5 else "0"
                    break
            
            if type_ref:
                init_value = etree.SubElement(spec, f"{{{self.NS}}}INIT-VALUE")
                
                if type_ref in self.type_definitions:
                    cat, sub_defs = self.type_definitions[type_ref]
                    
                    if cat == "struct":
                        # 结构体类型使用 RecordValueSpecification
                        record_value = etree.SubElement(init_value, f"{{{self.NS}}}RECORD-VALUE-SPECIFICATION")
                        etree.SubElement(record_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{type_ref}_Init"
                        
                        # 为每个字段添加初始值
                        if sub_defs:
                            fields_elem = etree.SubElement(record_value, f"{{{self.NS}}}FIELDS")
                            for field_name, field_type in sub_defs:
                                # 根据字段类型设置初始值
                                if field_type in self.type_definitions and self.type_definitions[field_type][0] == "struct":
                                    # 嵌套结构体 - 使用RECORD-VALUE-SPECIFICATION
                                    nested_record = etree.SubElement(fields_elem, f"{{{self.NS}}}RECORD-VALUE-SPECIFICATION")
                                    etree.SubElement(nested_record, f"{{{self.NS}}}SHORT-LABEL").text = f"{field_name}_Init"
                                    # 为嵌套结构体的所有字段添加初始值为0
                                    nested_defs = self.type_definitions[field_type][1]
                                    if nested_defs:
                                        nested_fields = etree.SubElement(nested_record, f"{{{self.NS}}}FIELDS")
                                        for nested_field_name, nested_field_type in nested_defs:
                                            nested_field_elem = etree.SubElement(nested_fields, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                                            etree.SubElement(nested_field_elem, f"{{{self.NS}}}SHORT-LABEL").text = f"{nested_field_name}_Init"
                                            etree.SubElement(nested_field_elem, f"{{{self.NS}}}VALUE").text = "0"
                                elif field_type in self.BASIC_TYPES or field_type in ['boolean', 'uint8', 'uint16', 'uint32', 'uint64', 'sint8', 'sint16', 'sint32', 'sint64', 'float32', 'float64']:
                                    # 基本类型字段 - 直接使用NUMERICAL-VALUE-SPECIFICATION
                                    num_value = etree.SubElement(fields_elem, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                                    etree.SubElement(num_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{field_name}_Init"
                                    etree.SubElement(num_value, f"{{{self.NS}}}VALUE").text = init_val if init_val != "{}" else "0"
                                else:
                                    # 其他类型字段 - 直接使用NUMERICAL-VALUE-SPECIFICATION
                                    num_value = etree.SubElement(fields_elem, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                                    etree.SubElement(num_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{field_name}_Init"
                                    etree.SubElement(num_value, f"{{{self.NS}}}VALUE").text = "0"
                    
                    elif cat == "array":
                        # 数组类型使用 ArrayValueSpecification
                        array_value = etree.SubElement(init_value, f"{{{self.NS}}}ARRAY-VALUE-SPECIFICATION")
                        etree.SubElement(array_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{type_ref}_Init"
                        
                        # 为数组元素添加初始值
                        elements_elem = etree.SubElement(array_value, f"{{{self.NS}}}ELEMENTS")
                        
                        # 获取数组第一个定义来解析数组大小
                        array_size = None
                        element_type = None
                        
                        if sub_defs and len(sub_defs) > 0:
                            first_def = sub_defs[0] if isinstance(sub_defs[0], str) else sub_defs[0][1] if isinstance(sub_defs[0], tuple) else sub_defs[0]
                            array_match = self.ARRAY_PATTERN.match(first_def)
                            if array_match:
                                element_type = array_match.group(1)
                                array_size = int(array_match.group(2))
                        
                        if array_size is None:
                            array_match = self.ARRAY_PATTERN.match(type_ref)
                            if array_match:
                                element_type = array_match.group(1)
                                array_size = int(array_match.group(2))
                        
                        # 生成数组元素初始值
                        if array_size and element_type:
                            for i in range(array_size):
                                elem_value = etree.SubElement(elements_elem, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                                etree.SubElement(elem_value, f"{{{self.NS}}}SHORT-LABEL").text = f"{element_type}_{i}_Init"
                                default_val = "0"
                                if init_val and init_val != "[]" and init_val.isdigit():
                                    default_val = init_val
                                etree.SubElement(elem_value, f"{{{self.NS}}}VALUE").text = default_val
                    
                    else:
                        # 基本类型使用 NumericalValueSpecification
                        num_value = etree.SubElement(init_value, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                        etree.SubElement(num_value, f"{{{self.NS}}}VALUE").text = init_val if init_val else "0"
                else:
                    # 不在type_definitions中的类型（如基本类型uint16, boolean）也使用NumericalValueSpecification
                    num_value = etree.SubElement(init_value, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                    etree.SubElement(num_value, f"{{{self.NS}}}VALUE").text = init_val if init_val else "0"
            
            # 为 R-PORT 添加接口引用
            iface_ref = etree.SubElement(port, f"{{{self.NS}}}REQUIRED-INTERFACE-TREF")
            iface_ref.set("DEST", "SENDER-RECEIVER-INTERFACE")
            iface_ref.text = f"/InterfaceTypes/{iface_name}"
        
        return port

    def _create_port(self, name, iface_name, direction):
        tag = "R-PORT-PROTOTYPE" if direction == 'R' else "P-PORT-PROTOTYPE"
        port = etree.Element(f"{{{self.NS}}}{tag}")
        port.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(port, f"{{{self.NS}}}SHORT-NAME").text = name
        com_specs = etree.SubElement(port, f"{{{self.NS}}}{'REQUIRED' if direction == 'R' else 'PROVIDED'}-COM-SPECS")
        spec_tag = "NONQUEUED-RECEIVER-COM-SPEC" if direction == 'R' else "NONQUEUED-SENDER-COM-SPEC"
        spec = etree.SubElement(com_specs, f"{{{self.NS}}}{spec_tag}")
        spec.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        deref = etree.SubElement(spec, f"{{{self.NS}}}DATA-ELEMENT-REF")
        deref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
        deref.text = f"/Interfaces/{iface_name}/{iface_name}"
        etree.SubElement(spec, f"{{{self.NS}}}HANDLE-OUT-OF-RANGE").text = "NONE"
        etree.SubElement(spec, f"{{{self.NS}}}USES-END-TO-END-PROTECTION").text = "false"
        etree.SubElement(spec, f"{{{self.NS}}}ALIVE-TIMEOUT").text = "0"
        return port

    def _create_runnable_with_port_access(self, runnable_name: str, task_name: str, port_accesses: List[Tuple[str, str, str]]):
        """创建包含端口访问的 Runnable"""
        runnable = etree.Element(f"{{{self.NS}}}RUNNABLE-ENTITY")
        runnable.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(runnable, f"{{{self.NS}}}SHORT-NAME").text = runnable_name
        
        # 添加最小启动间隔
        min_start_interval = etree.SubElement(runnable, f"{{{self.NS}}}MINIMUM-START-INTERVAL")
        min_start_interval.text = "0"
        
        # 添加并发调用信息
        concurrent = etree.SubElement(runnable, f"{{{self.NS}}}CAN-BE-INVOKED-CONCURRENTLY")
        concurrent.text = "false"
        
        # 添加端口访问点
        for port_name, port_direction, interface_name in port_accesses:
            if port_direction == 'R':  # 接收端口 - 数据接收点
                receive_points = etree.SubElement(runnable, f"{{{self.NS}}}DATA-RECEIVE-POINT-BY-ARGUMENTS")
                var_access = etree.SubElement(receive_points, f"{{{self.NS}}}VARIABLE-ACCESS")
                var_access.set("UUID", self._generate_uuid())
                
                # 生成变量访问名称
                access_name = f"IN_{port_name}_{interface_name}"
                etree.SubElement(var_access, f"{{{self.NS}}}SHORT-NAME").text = access_name
                
                accessed_var = etree.SubElement(var_access, f"{{{self.NS}}}ACCESSED-VARIABLE")
                autosar_var = etree.SubElement(accessed_var, f"{{{self.NS}}}AUTOSAR-VARIABLE-IREF")
                
                port_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}PORT-PROTOTYPE-REF")
                port_ref.set("DEST", "R-PORT-PROTOTYPE")
                port_ref.text = f"/Components/{task_name.split('_')[0] if '_' in task_name else task_name}/{port_name}"
                
                target_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}TARGET-DATA-PROTOTYPE-REF")
                target_ref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
                target_ref.text = f"/Interfaces/{interface_name}/{interface_name}"
                
            elif port_direction == 'S':  # 发送端口 - 数据发送点
                send_points = etree.SubElement(runnable, f"{{{self.NS}}}DATA-SEND-POINT-BY-ARGUMENTS")
                var_access = etree.SubElement(send_points, f"{{{self.NS}}}VARIABLE-ACCESS")
                var_access.set("UUID", self._generate_uuid())
                
                # 生成变量访问名称
                access_name = f"OUT_{port_name}_{interface_name}"
                etree.SubElement(var_access, f"{{{self.NS}}}SHORT-NAME").text = access_name
                
                accessed_var = etree.SubElement(var_access, f"{{{self.NS}}}ACCESSED-VARIABLE")
                autosar_var = etree.SubElement(accessed_var, f"{{{self.NS}}}AUTOSAR-VARIABLE-IREF")
                
                port_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}PORT-PROTOTYPE-REF")
                port_ref.set("DEST", "P-PORT-PROTOTYPE")
                port_ref.text = f"/Components/{task_name.split('_')[0] if '_' in task_name else task_name}/{port_name}"
                
                target_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}TARGET-DATA-PROTOTYPE-REF")
                target_ref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
                target_ref.text = f"/Interfaces/{interface_name}/{interface_name}"
        
        return runnable

    def _get_current_time_string(self):
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _validate_types(self):
        """验证所有引用的类型都存在（基本类型和自定义类型都要检查）"""
        errors = []
        
        # 检查 API 条目中引用的类型
        for api_name, type_ref, port_dir, swc_name, task, init_val in self.api_entries:
            if not self._type_exists(type_ref):
                errors.append(f"❌ API '{api_name}' 引用的类型不存在: {type_ref}")
        
        # 检查结构体字段引用的类型
        for type_name, (cat, defs) in self.type_definitions.items():
            if cat == "struct" and defs:
                for field_name, field_type in defs:
                    if not self._type_exists(field_type):
                        errors.append(f"❌ 结构体 '{type_name}' 的字段 '{field_name}' 引用的类型不存在: {field_type}")
            elif cat == "array" and defs:
                # 数组类型检查元素类型
                first_def = defs[0] if isinstance(defs[0], str) else defs[0][1] if isinstance(defs[0], tuple) else defs[0]
                match = self.ARRAY_PATTERN.match(first_def)
                if match:
                    element_type = match.group(1)
                    if not self._type_exists(element_type):
                        errors.append(f"❌ 数组类型 '{type_name}' 的元素类型不存在: {element_type}")
        
        if errors:
            print("\n" + "="*60)
            print("❌ 类型验证失败！")
            print("="*60)
            for error in errors:
                print(error)
            print("="*60)
            print("\n请检查并修正 Excel 中的类型定义。")
            exit(1)
        
        print("✅ 所有类型验证通过")

    def _type_exists(self, type_name: str) -> bool:
        """检查类型是否存在（包括基本类型和自定义类型）"""
        # 检查是否是基本类型
        if type_name in self.BASIC_TYPES or type_name in self.platform_basic_types:
            return True
        
        # 检查是否是自定义类型
        if type_name in self.type_definitions:
            return True
        
        # 检查是否是数组格式的引用（如 uint8[10]）
        if self.ARRAY_PATTERN.match(type_name):
            match = self.ARRAY_PATTERN.match(type_name)
            if match:
                element_type = match.group(1)
                return self._type_exists(element_type)
        
        return False

    # === Main Generate ===
    def generate(self):
        self._read_type_definitions()
        self._read_api_definitions()
        
        # 验证所有类型
        self._validate_types()

        nsmap = {
            None: self.NS,
            'xsi': 'http://www.w3.org/2001/XMLSchema-instance'
        }
        root = etree.Element(f"{{{self.NS}}}AUTOSAR", nsmap=nsmap)
        root.set("{http://www.w3.org/2001/XMLSchema-instance}schemaLocation",
                 f"{self.NS} AUTOSAR_00048.xsd")

        comment = etree.Comment("support: Atech Co. tool version: 1.0.1 export time: " + 
                               self._get_current_time_string())
        root.addprevious(comment)

        ar_pkgs = root.find(f"{{{self.NS}}}AR-PACKAGES")
        if ar_pkgs is None:
            ar_pkgs = etree.SubElement(root, f"{{{self.NS}}}AR-PACKAGES")

        # 根据 sample.arxml 的格式，我们需要创建类似的结构
        # 1. 创建 Atomic 包（包含 SWC）
        atomic_pkg = etree.SubElement(ar_pkgs, f"{{{self.NS}}}AR-PACKAGE")
        atomic_pkg.set("UUID", self._generate_uuid())
        etree.SubElement(atomic_pkg, f"{{{self.NS}}}SHORT-NAME").text = "Atomic"
        atomic_elems = etree.SubElement(atomic_pkg, f"{{{self.NS}}}ELEMENTS")

        # 2. 创建 InterfaceTypes 包（包含接口）
        interface_pkg = etree.SubElement(ar_pkgs, f"{{{self.NS}}}AR-PACKAGE")
        interface_pkg.set("UUID", self._generate_uuid())
        etree.SubElement(interface_pkg, f"{{{self.NS}}}SHORT-NAME").text = "InterfaceTypes"
        interface_elems = etree.SubElement(interface_pkg, f"{{{self.NS}}}ELEMENTS")

        # 3. 创建本地数据类型包（从配置中获取包名）
        local_pkg_name = self.platform_base_path.split('/')[-1] if self.platform_base_path else "DataTypeFromSignalOrGroup"
        datatype_pkg = etree.SubElement(ar_pkgs, f"{{{self.NS}}}AR-PACKAGE")
        datatype_pkg.set("UUID", self._generate_uuid())
        etree.SubElement(datatype_pkg, f"{{{self.NS}}}SHORT-NAME").text = local_pkg_name
        datatype_elems = etree.SubElement(datatype_pkg, f"{{{self.NS}}}ELEMENTS")

        # 创建 Implementation Data Types - 按照原始 Excel 中的顺序
        self.created_idt_types.clear()
        
        # 首先，收集所有需要的类型
        all_types_needed = set(entry[1] for entry in self.api_entries)
        
        # 按照原始 Excel 中的类型顺序创建
        for tname in self.type_order:
            if tname in all_types_needed and tname not in self.created_idt_types:
                self._create_and_add_type(tname, datatype_elems)
        
        # 然后创建其他需要的类型（不在原始类型定义中但被引用的类型）
        for tname in all_types_needed:
            if tname not in self.created_idt_types:
                self._create_and_add_type(tname, datatype_elems)
        
        # 确保所有类型都按照原始顺序排列
        # 重新排序 datatype_elems 中的元素
        type_elements = list(datatype_elems)
        datatype_elems.clear()
        
        # 按照 type_order 的顺序重新添加元素
        for tname in self.type_order:
            for elem in type_elements:
                name_elem = elem.find(f"{{{self.NS}}}SHORT-NAME")
                if name_elem is not None and name_elem.text == tname:
                    datatype_elems.append(elem)
                    break

        # 创建 Sender-Receiver Interfaces
        created_ifaces = set()
        for entry in self.api_entries:
            api_name, type_ref, _, _, _, _ = entry
            if api_name not in created_ifaces:
                iface = self._create_interface_correct_format(api_name, type_ref)
                interface_elems.append(iface)
                created_ifaces.add(api_name)

        # SWC Components - 按 SWC 分组
        swc_groups: Dict[str, List] = {}
        for entry in self.api_entries:
            _, _, _, swc, _, _ = entry
            swc_groups.setdefault(swc, []).append(entry)

        for swc_name, entries in swc_groups.items():
            # 创建 SWC 组件
            comp = etree.SubElement(atomic_elems, f"{{{self.NS}}}APPLICATION-SW-COMPONENT-TYPE")
            comp.set("UUID", self._generate_uuid())
            etree.SubElement(comp, f"{{{self.NS}}}SHORT-NAME").text = swc_name

            ports = etree.SubElement(comp, f"{{{self.NS}}}PORTS")
            for api_name, _, port_dir, _, _, _ in entries:
                ports.append(self._create_port_correct_format(api_name, api_name, port_dir))

            # 创建内部行为
            internal = etree.SubElement(comp, f"{{{self.NS}}}INTERNAL-BEHAVIORS")
            ib = etree.SubElement(internal, f"{{{self.NS}}}SWC-INTERNAL-BEHAVIOR")
            ib.set("UUID", self._generate_uuid())
            etree.SubElement(ib, f"{{{self.NS}}}SHORT-NAME").text = "default"

            # 按任务分组
            task_groups = {}
            for api_name, type_ref, port_dir, _, task, _ in entries:
                if task not in task_groups:
                    task_groups[task] = []
                task_groups[task].append((api_name, port_dir))
            
            # 为每个任务创建事件和 Runnable
            events = etree.SubElement(ib, f"{{{self.NS}}}EVENTS")
            runnables = etree.SubElement(ib, f"{{{self.NS}}}RUNNABLES")
            
            for task_name, task_entries in task_groups.items():
                # 创建事件
                timing_event = etree.SubElement(events, f"{{{self.NS}}}TIMING-EVENT")
                timing_event.set("UUID", self._generate_uuid())
                etree.SubElement(timing_event, f"{{{self.NS}}}SHORT-NAME").text = task_name
                
                # 创建 Runnable（使用任务名作为 Runnable 名）
                runnable_name = task_name
                runnable = etree.SubElement(runnables, f"{{{self.NS}}}RUNNABLE-ENTITY")
                runnable.set("UUID", self._generate_uuid())
                etree.SubElement(runnable, f"{{{self.NS}}}SHORT-NAME").text = runnable_name
                etree.SubElement(runnable, f"{{{self.NS}}}MINIMUM-START-INTERVAL").text = "0.0"
                etree.SubElement(runnable, f"{{{self.NS}}}CAN-BE-INVOKED-CONCURRENTLY").text = "false"
                
                # 为 Runnable 添加端口访问点
                for api_name, port_dir in task_entries:
                    if port_dir == 'R':  # 接收端口 - 使用 DATA-READ-ACCESSS
                        read_access = etree.SubElement(runnable, f"{{{self.NS}}}DATA-READ-ACCESSS")
                        var_access = etree.SubElement(read_access, f"{{{self.NS}}}VARIABLE-ACCESS")
                        var_access.set("UUID", self._generate_uuid())
                        etree.SubElement(var_access, f"{{{self.NS}}}SHORT-NAME").text = api_name  # 使用原始 API 名称，不加前缀
                        
                        accessed_var = etree.SubElement(var_access, f"{{{self.NS}}}ACCESSED-VARIABLE")
                        autosar_var = etree.SubElement(accessed_var, f"{{{self.NS}}}AUTOSAR-VARIABLE-IREF")
                        
                        port_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}PORT-PROTOTYPE-REF")
                        port_ref.set("DEST", "R-PORT-PROTOTYPE")
                        port_ref.text = f"/Atomic/{swc_name}/{api_name}"
                        
                        target_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}TARGET-DATA-PROTOTYPE-REF")
                        target_ref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
                        target_ref.text = f"/InterfaceTypes/{api_name}/{api_name}"
                        
                    elif port_dir == 'S':  # 发送端口 - 使用 DATA-WRITE-ACCESSS
                        write_access = etree.SubElement(runnable, f"{{{self.NS}}}DATA-WRITE-ACCESSS")
                        var_access = etree.SubElement(write_access, f"{{{self.NS}}}VARIABLE-ACCESS")
                        var_access.set("UUID", self._generate_uuid())
                        etree.SubElement(var_access, f"{{{self.NS}}}SHORT-NAME").text = api_name  # 使用原始 API 名称，不加前缀
                        
                        accessed_var = etree.SubElement(var_access, f"{{{self.NS}}}ACCESSED-VARIABLE")
                        autosar_var = etree.SubElement(accessed_var, f"{{{self.NS}}}AUTOSAR-VARIABLE-IREF")
                        
                        port_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}PORT-PROTOTYPE-REF")
                        port_ref.set("DEST", "P-PORT-PROTOTYPE")
                        port_ref.text = f"/Atomic/{swc_name}/{api_name}"
                        
                        target_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}TARGET-DATA-PROTOTYPE-REF")
                        target_ref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
                        target_ref.text = f"/InterfaceTypes/{api_name}/{api_name}"
                
                # 设置事件引用
                start_on_event_ref = etree.SubElement(timing_event, f"{{{self.NS}}}START-ON-EVENT-REF")
                start_on_event_ref.set("DEST", "RUNNABLE-ENTITY")
                start_on_event_ref.text = f"/Atomic/{swc_name}/default/{runnable_name}"
                
                etree.SubElement(timing_event, f"{{{self.NS}}}PERIOD").text = "0.01"
                etree.SubElement(runnable, f"{{{self.NS}}}SYMBOL").text = runnable_name
            
            etree.SubElement(ib, f"{{{self.NS}}}SUPPORTS-MULTIPLE-INSTANTIATION").text = "false"

        # 写入文件，使用正确的 XML 声明格式
        xml_content = etree.tostring(root, pretty_print=True, encoding='UTF-8', xml_declaration=True)
        # 修复 XML 声明格式（使用双引号）
        xml_content = xml_content.replace(b"<?xml version='1.0' encoding='UTF-8'?>", b'<?xml version="1.0" encoding="UTF-8"?>')
        
        with open(self.output_arxml, 'wb') as f:
            f.write(xml_content)
        print(f"✅ 生成完成！包含 {len(swc_groups)} 个 SWC，输出: {self.output_arxml}")


# === 主程序入口 ===
if __name__ == '__main__':
    # 支持命令行参数，默认使用 'converted_from_arxml.xlsx' 和 'SwcOutput.arxml'
    if len(sys.argv) >= 2:
        input_excel = sys.argv[1]
        output_arxml = sys.argv[2] if len(sys.argv) >= 3 else 'SwcOutput.arxml'
    else:
        input_excel = 'converted_from_arxml.xlsx'
        output_arxml = 'SwcOutput.arxml'

    if not os.path.exists(input_excel):
        print(f"❌ 输入文件不存在: {input_excel}")
        print("请确保 Excel 文件存在，并包含 'Types' 和 'APIs' 工作表。")
        exit(1)

    generator = SwcTaskArxmlGenerator(input_excel, output_arxml)
    generator.generate()